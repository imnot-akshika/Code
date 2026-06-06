import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
from pathlib import Path
from collections import defaultdict
from pypdf import PdfReader, PdfWriter


class ReportGenerator:
    def __init__(self, csv_file: str):
        self.data = self._load_csv(csv_file)

    def _load_csv(self, filename: str) -> list[dict]:
        with open(filename) as f:
            return list(csv.DictReader(f))

    def generate_excel(self, output_file: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # write bold headers
        headers = ["Name", "Sales", "Region"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # write data rows
        for row_idx, record in enumerate(self.data, 2):
            ws.cell(row=row_idx, column=1, value=record["name"])
            ws.cell(row=row_idx, column=2, value=float(record["sales"]))
            ws.cell(row=row_idx, column=3, value=record["region"])

        # auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

        wb.save(output_file)
        print(f"Excel report saved to {output_file}")

    def summary_by_region(self) -> dict[str, float]:
        result = defaultdict(float)
        for row in self.data:
            result[row["region"]] += float(row["sales"])
        return dict(result)

    def generate_text_summary(self, output_file: str) -> None:
        total_sales = sum(float(row["sales"]) for row in self.data)
        top_performer = max(self.data, key=lambda x: float(x["sales"]))
        region_total = self.summary_by_region()

        with open(output_file, "w") as f:
            f.write(f"=== Sales Summary ===\n\n")
            f.write(f"Total Sales: ${total_sales:.2f}\n")
            f.write(f"Top Performer: {top_performer['name']} with ${float(top_performer['sales']):.2f}\n\n")
            f.write("Sales by Region:\n")
            for region, total in region_total.items():
                f.write(f"  {region}: ${total:.2f}\n")
        
        print(f"Text summary saved to {output_file}")





#example usage
import csv

rows = [
    ["name", "sales", "region"],
    ["Alice", "15000", "North"],
    ["Bob", "12000", "South"],
    ["Charlie", "18000", "North"],
    ["Diana", "9500", "South"],
    ["Eve", "21000", "East"],
]
with open("sales.csv", "w", newline="") as f:
    csv.writer(f).writerows(rows)

rg = ReportGenerator("sales.csv")
rg.generate_excel("report.xlsx")
print(rg.summary_by_region())
rg.generate_text_summary("summary.txt")